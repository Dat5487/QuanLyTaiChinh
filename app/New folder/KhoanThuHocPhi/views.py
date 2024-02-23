from django.shortcuts import render, redirect
from django.urls import reverse
from ThongTin.models import HeDaoTao, NganhDaoTao, Khoa, KhoaHoc, Lop
from ThongTin.models import SinhVien
from KhoanThuHocPhi.models import KhoanThuHocPhi
from ThietLapMucHocPhi.models import ThietLapMucHocPhi
from openpyxl import Workbook
from django.shortcuts import get_object_or_404
from .forms import KhoanThuHocPhiForm, BatchKhoanThuHocPhiForm
from django.http import HttpResponse
from datetime import datetime
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required

@login_required
def ListKhoanThuHocPhi(request):
    ListHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    ListNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    ListKhoa = Khoa.objects.all().order_by('-id')
    ListKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    ListLop = Lop.objects.all().order_by('-id')
    if request.method == 'POST':
        loaiBaoCao = request.POST.get('loaiBaoCao')
        if loaiBaoCao == 'tongHop':
            BaoCaoTongHop()

    return render(request, 'listKhoanThu.html', {'ListHeDaoTao': ListHeDaoTao, 'ListNganhDaoTao': ListNganhDaoTao, 'ListKhoa': ListKhoa, 'ListKhoaHoc': ListKhoaHoc, 'ListLop': ListLop})

@login_required
def ListSinhVien(request,id):
   lop = get_object_or_404(Lop, id=id)
   ListSinhVien = SinhVien.objects.filter(lop=id).order_by('-id')
   listKhoanThuHocPhi = KhoanThuHocPhi.objects.filter(mucHocPhi__lop = lop.lop).order_by('-id')
   alertContent = request.GET.get('alertContent')
   return render(request, 'listSinhVien.html', {'listKhoanThuHocPhi': listKhoanThuHocPhi,'ListSinhVien': ListSinhVien, 'lop': lop,'alertContent': alertContent,'lopId':id})

#Batch insert 
@login_required
def AddKhoanThuHocPhi(request,id):
    lop = get_object_or_404(Lop, id=id)
    listSinhVien = SinhVien.objects.filter(lop__id = id).order_by('-id')
    listMucHocPhi= ThietLapMucHocPhi.objects.filter(lop = lop.lop).order_by('-id')
    if request.method == 'POST':
        form = BatchKhoanThuHocPhiForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTienMienGiam'] = data_copy['soTienMienGiam'].replace(",", "") 
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            mucHocPhiId = request.POST.get('mucHocPhi')
            mucHocPhi = get_object_or_404(ThietLapMucHocPhi, id=mucHocPhiId)
            hanNop  = request.POST.get('hanNop')
            noiDungThu  = request.POST.get('noiDungThu')
            ghiChu = request.POST.get('ghiChu')
            maSoThue = request.POST.get('maSoThue')
            soTienMienGiam = request.POST.get('soTienMienGiam').replace(",", "") 
            soTien = request.POST.get('soTien').replace(",", "") 
            for sinhVien in listSinhVien:
                khoanThuHocPhi = KhoanThuHocPhi(sinhVien=sinhVien, mucHocPhi=mucHocPhi, hanNop = hanNop , noiDungThu=noiDungThu, ghiChu=ghiChu,
                                                maSoThue = maSoThue, soTienMienGiam = soTienMienGiam, soTien = soTien, trangThai = 0 )
                khoanThuHocPhi.save()
            alert_content = 'Thêm'
            url = reverse('listSinhVien', args=[id]) + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = BatchKhoanThuHocPhiForm()
    return render(request, 'addKhoanThuHocPhi.html', {'form': form, 'listSinhVien': listSinhVien,'listMucHocPhi': listMucHocPhi,'lopId':id})

@login_required
def NewKhoanThuHocPhi(request,id):
    lop = get_object_or_404(Lop, id=id)
    listSinhVien = SinhVien.objects.filter(lop__id = id).order_by('-id')
    listMucHocPhi= ThietLapMucHocPhi.objects.filter(lop = lop.lop).order_by('-id')
    if request.method == 'POST':
        form = KhoanThuHocPhiForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTienMienGiam'] = data_copy['soTienMienGiam'].replace(",", "") 
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listSinhVien', args=[id]) + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = KhoanThuHocPhiForm()
    return render(request, 'newKhoanThuHocPhi.html', {'form': form, 'listSinhVien': listSinhVien,'listMucHocPhi': listMucHocPhi,'lopId':id})


@login_required
def SuaKhoanThuHocPhi(request, id):
    
    khoanThuHocPhi = KhoanThuHocPhi.objects.get(id=id)
    listMucHocPhi= ThietLapMucHocPhi.objects.filter(lop = khoanThuHocPhi.sinhVien.lop.lop).order_by('-id')
    if request.method == 'POST':
        khoanThuHocPhi.sinhVien = get_object_or_404(SinhVien, id=request.POST.get('sinhVien'))
        khoanThuHocPhi.mucHocPhi = get_object_or_404(ThietLapMucHocPhi, id=request.POST.get('mucHocPhi'))
        khoanThuHocPhi.hanNop = request.POST.get('hanNop')
        khoanThuHocPhi.noiDungThu = request.POST.get('noiDungThu')
        khoanThuHocPhi.ghiChu = request.POST.get('ghiChu')
        khoanThuHocPhi.maSoThue = request.POST.get('maSoThue')
        khoanThuHocPhi.soTienMienGiam = request.POST.get('soTienMienGiam').replace(",", "")
        khoanThuHocPhi.soTien = request.POST.get('soTien').replace(",", "")

        khoanThuHocPhi.save()
        alert_content = 'Sửa'
        url = reverse('listSinhVien', args=[khoanThuHocPhi.sinhVien.lop.id]) + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        
        form = KhoanThuHocPhiForm()

    return render(request, 'suaKhoanThuHocPhi.html', {'form': form,'khoanThuHocPhi': khoanThuHocPhi,'listMucHocPhi': listMucHocPhi})

@login_required
def XoaKhoanThuHocPhi(request, id):
    khoanThuHocPhi = get_object_or_404(KhoanThuHocPhi, id=id)

    if request.method == 'POST':
        khoanThuHocPhi.delete()
        alert_content = 'Xóa'
        url = reverse('listSinhVien', args=[khoanThuHocPhi.sinhVien.lop.id]) + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'xoaKhoanThuHocPhi.html', {'khoanThuHocPhi': khoanThuHocPhi})

@login_required
def XemKhoanThuHocPhi(request, id):
    khoanThuHocPhi = get_object_or_404(KhoanThuHocPhi, id=id)
    tongSoTien = format(khoanThuHocPhi.soTien, ",.0f")
    return render(request, 'xemKhoanThuHocPhi.html', {'khoanThuHocPhi': khoanThuHocPhi})

@login_required
def XemCacKhoanThuCuaSV(request,id):
   sinhVien = get_object_or_404(SinhVien, id=id)
   listKhoanThu = KhoanThuHocPhi.objects.filter(sinhVien=id).order_by('-id')
   alertContent = request.GET.get('alertContent')
   return render(request, 'xemCacKhoanThuCuaSV.html', {'listKhoanThu': listKhoanThu,'sinhVien': sinhVien,'alertContent': alertContent})

def BaoCaoTongHop(self):
    ListKhoa = Khoa.objects.all().order_by('-id')
    filename = 'data_export.xlsx'

    # Create a new workbook and get the active sheet
    workbook = Workbook()
    worksheet = workbook.active
    for khoa in ListKhoa:
        queryset = KhoanThuHocPhi.objects.filter(mucHocPhi__khoa=khoa.tenKhoa)
        worksheet = workbook.create_sheet(title=khoa.tenKhoa)
        # Write column headers
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name

            column_letter = get_column_letter(col_num)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = max(column_dimensions.width, len(column_name) + 2)

        # Write data rows
        for row_num, obj in enumerate(queryset, 2):
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                format(obj.mucHocPhi.soTien, ","),
                format(obj.soTienMienGiam, ","),
                format(obj.soTien, ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                
                # Remove time zone information from datetime objects
                if isinstance(value, datetime):
                    value = value.replace(tzinfo=None)
                
                cell.value = value
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = max(column_dimensions.width, len(str(value)), len(column_names[col_num-1])) + 2
    
    # Set response headers for file download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    # Save the workbook to the response
    workbook.save(response)
    return response