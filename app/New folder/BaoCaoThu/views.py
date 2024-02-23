from django.shortcuts import render, redirect
from django.urls import reverse
from ThongTin.models import HeDaoTao, NganhDaoTao, Khoa, KhoaHoc, Lop, NamHoc
from ThongTin.models import SinhVien
from KhoanThuHocPhi.models import KhoanThuHocPhi
from ThietLapMucHocPhi.models import ThietLapMucHocPhi
from django_excel import make_response
from openpyxl import Workbook
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from datetime import datetime
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.db.models import Q
import urllib.parse
from django.db.models import Sum

@login_required
def ChonBaoCao(request):
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacKhoa = Khoa.objects.all().order_by('-id')
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacNamHoc = NamHoc.objects.all().order_by('-id')
    cacLop = Lop.objects.all().order_by('-id')
    cacLoaiBaoCao = ["Báo cáo tổng hợp","Báo cáo tổng hợp theo lớp","Báo cáo công nợ sinh viên","Báo cáo công nợ sinh viên theo lớp"]

    if request.method == 'POST':
        loaiBaoCao = request.POST.get('loaiBaoCao')
        if loaiBaoCao == 'Báo cáo tổng hợp':
            return BaoCaoTongHop()
        elif loaiBaoCao == 'Báo cáo tổng hợp theo lớp':
            lop = request.POST.get('lop')
            return BaoCaoTongHopTheoLop(lop)
        elif loaiBaoCao == 'Báo cáo công nợ sinh viên':
            return BaoCaoCongNoSinhVien()
        elif loaiBaoCao == 'Báo cáo công nợ sinh viên theo lớp':
            lop = request.POST.get('lop')
            return BaoCaoCongNoTheoLop(lop)

    return render(request, 'chonBaoCao.html', {'cacHeDaoTao': cacHeDaoTao, 'cacNganhDaoTao': cacNganhDaoTao, 'cacKhoa': cacKhoa, 'cacNamHoc': cacNamHoc, 'cacKhoaHoc': cacKhoaHoc, 'cacLop': cacLop, 'cacLoaiBaoCao': cacLoaiBaoCao})

def BaoCaoTongHop():
    ListKhoa = Khoa.objects.all().order_by('-id')
    workbook = Workbook()
    worksheet = workbook.active
    for khoa in ListKhoa:
        queryset = KhoanThuHocPhi.objects.filter(mucHocPhi__khoa=khoa.tenKhoa).order_by('sinhVien__maSV')
        worksheet = workbook.create_sheet(title=khoa.tenKhoa)
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if obj.mucHocPhi.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                obj.mucHocPhi.namHoc,
                hocKy,
                format(obj.mucHocPhi.soTien, ","),
                format(obj.soTienMienGiam, ","),
                format(obj.soTien, ","),
                trangThai

            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = urllib.parse.quote('Báo cáo tổng hợp.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response


def BaoCaoTongHopTheoLop(lopId):
    lopDuocChon = get_object_or_404(Lop, id=lopId)
    listNamHoc = NamHoc.objects.all().order_by('-id')
    name = 'Báo cáo tổng hợp theo lớp '+ lopDuocChon.lop +' .xlsx'
    filename = urllib.parse.quote(name)
    workbook = Workbook()
    worksheet = workbook.active
    for namHoc in listNamHoc:
        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=1) & Q(mucHocPhi__lop=lopDuocChon.lop))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 1")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if obj.mucHocPhi.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                obj.mucHocPhi.namHoc,
                hocKy,
                format(obj.mucHocPhi.soTien, ","),
                format(obj.soTienMienGiam, ","),
                format(obj.soTien, ","),
                trangThai
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=0) & Q(mucHocPhi__lop=lopDuocChon.lop))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 2")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if obj.mucHocPhi.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                obj.mucHocPhi.namHoc,
                hocKy,
                format(obj.mucHocPhi.soTien, ","),
                format(obj.soTienMienGiam, ","),
                format(obj.soTien, ","),
                trangThai
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

def BaoCaoCongNoSinhVien():
    ListKhoa = Khoa.objects.all().order_by('-id')
    workbook = Workbook()
    worksheet = workbook.active
    for khoa in ListKhoa:
        queryset = KhoanThuHocPhi.objects.filter(mucHocPhi__khoa=khoa.tenKhoa).filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('mucHocPhi__soTien'))
        worksheet = workbook.create_sheet(title=khoa.tenKhoa)
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']

        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = urllib.parse.quote('Báo cáo công nợ sinh viên.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

def BaoCaoCongNoTheoLop(lopId):
    lopDuocChon = get_object_or_404(Lop, id=lopId)
    listNamHoc = NamHoc.objects.all().order_by('-id')
    name = 'Báo cáo công nợ sinh viên theo lớp '+ lopDuocChon.lop +' .xlsx'
    filename = urllib.parse.quote(name)
    workbook = Workbook()
    worksheet = workbook.active
    for namHoc in listNamHoc:
        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=1) & Q(mucHocPhi__lop=lopDuocChon.lop)).filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('mucHocPhi__soTien'))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 1")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=0) & Q(mucHocPhi__lop=lopDuocChon.lop)).filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('mucHocPhi__soTien'))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 2")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response
