from django.shortcuts import render, redirect
from django.urls import reverse
from app.models import *
from app.forms import *
from app.vnpay import vnpay
from django_excel import make_response
from openpyxl import Workbook
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from datetime import datetime
from openpyxl.utils import get_column_letter
from django.contrib.auth.decorators import login_required
from django.db.models import Q
import urllib.parse
from django.db.models import Sum
from django.shortcuts import render, redirect, HttpResponse
from .forms import LoginForm
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.shortcuts import redirect


import hashlib
import hmac
import json
import urllib
import urllib.parse
import urllib.request
import random
import requests
from datetime import datetime
from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
from urllib.parse import quote


@login_required
def home(request):
    soLuongChuaNop = KhoanThuHocPhi.objects.filter(trangThai=0).count()
    soLuongDaNop = KhoanThuHocPhi.objects.filter(trangThai=1).count()
    soTienChuaNop = KhoanThuHocPhi.objects.filter(trangThai=0).aggregate(Sum('soTien'))['soTien__sum']
    soTienDaNop = KhoanThuHocPhi.objects.filter(trangThai=1).aggregate(Sum('soTien'))['soTien__sum']
    return render(request, 'home/home.html', {'soLuongChuaNop': soLuongChuaNop,'soLuongDaNop': soLuongDaNop,'soTienChuaNop': soTienChuaNop,'soTienDaNop': soTienDaNop })


@login_required
def ListMucHocPhi(request):
   alertContent = request.GET.get('alertContent')
   ListMucHocPhi = ThietLapMucHocPhi.objects.all().order_by('-ngayThem')
   return render(request, 'ThietLapMucHocPhi/listMucHocPhi.html', {'ListMucHocPhi': ListMucHocPhi, 'alertContent': alertContent})

@login_required
def NewMucHocPhi(request):
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    cacNamHoc = NamHoc.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacKhoa= Khoa.objects.all().order_by('-id')
    cacLop = Lop.objects.all().order_by('-id')
    if request.method == 'POST':
        form = ThietLapMucHocPhiForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        data_copy['isNganhHai'] = request.POST.get('isNganhHai')
        data_copy['isHocLai'] = request.POST.get('isHocLai')
        data_copy['isMienGiam'] = request.POST.get('isMienGiam')
        data_copy['isMonGDTC'] = request.POST.get('isMonGDTC')
        data_copy['isChatLuongCao'] = request.POST.get('isChatLuongCao')
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listMucHocPhi') + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = ThietLapMucHocPhiForm()
    return render(request, 'ThietLapMucHocPhi/newMucHocPhi.html', {'form': form, 'cacHeDaoTao': cacHeDaoTao,'cacNganhDaoTao': cacNganhDaoTao,'cacKhoa': cacKhoa,'cacLop': cacLop,'cacNamHoc': cacNamHoc})

@login_required
def SuaMucHocPhi(request, id):
    mucHocPhi = get_object_or_404(ThietLapMucHocPhi, id=id)
    soTien = format(mucHocPhi.soTien, ",.0f")

    if request.method == 'POST':
        mucHocPhi.heDaoTao = request.POST.get('heDaoTao')
        mucHocPhi.namHoc = request.POST.get('namHoc')
        mucHocPhi.tenNganh = request.POST.get('tenNganh')
        mucHocPhi.khoa = request.POST.get('khoa')
        mucHocPhi.lop = request.POST.get('lop')
        mucHocPhi.hocKy = request.POST.get('hocKy')
        mucHocPhi.isNganhHai = request.POST.get('isNganhHai')
        mucHocPhi.isHocLai = request.POST.get('isHocLai')
        mucHocPhi.isMienGiam = request.POST.get('isMienGiam')
        mucHocPhi.isChatLuongCao = request.POST.get('isChatLuongCao')
        mucHocPhi.isMonGDTC = request.POST.get('isMonGDTC')
        mucHocPhi.tinhChat = request.POST.get('tinhChat')
        mucHocPhi.heSo = request.POST.get('heSo')
        mucHocPhi.soTien = request.POST.get('soTien').replace(",", "")

        mucHocPhi.save()
        alert_content = 'Sửa'
        url = reverse('listMucHocPhi') + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        mucHocPhi = ThietLapMucHocPhi.objects.get(id=id)
        form = ThietLapMucHocPhiForm()

    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    cacNamHoc = NamHoc.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacKhoa= Khoa.objects.all().order_by('-id')
    cacLop = Lop.objects.all().order_by('-id')
    return render(request, 'ThietLapMucHocPhi/suaMucHocPhi.html', {'form': form,'mucHocPhi': mucHocPhi, 'cacHeDaoTao': cacHeDaoTao,'cacNamHoc': cacNamHoc,'cacNganhDaoTao': cacNganhDaoTao,'cacKhoa': cacKhoa,'cacLop': cacLop,'soTien': soTien})

@login_required
def XoaMucHocPhi(request, id):
    mucHocPhi = get_object_or_404(ThietLapMucHocPhi, id=id)
    soTien = format(mucHocPhi.soTien, ",.0f")
    if request.method == 'POST':
        mucHocPhi.delete()
        alert_content = 'Xóa'
        url = reverse('listMucHocPhi') + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'ThietLapMucHocPhi/xoaMucHocPhi.html', {'mucHocPhi': mucHocPhi,'soTien': soTien})
def login_page(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = request.POST["username"]
            password = request.POST["password"]
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect("/")
            else:
                return redirect("/login")
    else:
        form = LoginForm()
    return render(request, 'login/login.html', {'form': form})

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required
def ListKhoanThuNhapHoc(request):
   alertContent = request.GET.get('alertContent')
   listKhoanThuNhapHoc = KhoanThuNhapHoc.objects.all().order_by('-id')
   return render(request, 'KhoanThuNhapHoc/listKhoanThuNhapHoc.html', {'listKhoanThuNhapHoc': listKhoanThuNhapHoc, 'alertContent': alertContent})



@login_required
def NewKhoanThuNhapHoc(request):
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacTenNganh = NganhDaoTao.objects.all().order_by('-id')
    cacKhoanThu = DanhMucThuChi.objects.filter(isKhoanThuNhapHoc = True).order_by('-id')
    if request.method == 'POST':
        form = KhoanThuNhapHocForm(request.POST)
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listKhoanThuNhapHoc') + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = KhoanThuNhapHocForm()
    return render(request, 'KhoanThuNhapHoc/newKhoanThuNhapHoc.html', {'form': form, 'cacKhoaHoc': cacKhoaHoc,'cacTenNganh': cacTenNganh,'cacKhoanThu': cacKhoanThu})


@login_required
def SuaKhoanThuNhapHoc(request, id):
    khoanThuNhaphoc = get_object_or_404(KhoanThuNhapHoc, id=id)
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacTenNganh = NganhDaoTao.objects.all().order_by('-id')
    cacKhoanThu = DanhMucThuChi.objects.filter(isKhoanThuNhapHoc = True).order_by('-id')
    if request.method == 'POST':
        khoanThuNhaphoc.khoaHoc =  get_object_or_404(KhoaHoc, id=request.POST.get('khoaHoc'))
        khoanThuNhaphoc.dotNhapHoc = request.POST.get('dotNhapHoc')
        khoanThuNhaphoc.khoanThu = get_object_or_404(DanhMucThuChi, id=request.POST.get('khoanThu'))
        khoanThuNhaphoc.nganhDaoTao = get_object_or_404(NganhDaoTao, id=request.POST.get('nganhDaoTao'))
        khoanThuNhaphoc.soTien = request.POST.get('soTien')
        khoanThuNhaphoc.save()
        alert_content = 'Sửa'
        url = reverse('listKhoanThuNhapHoc') + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        khoanThuNhapHoc = KhoanThuNhapHoc.objects.get(id=id)
        form = KhoanThuNhapHocForm()


    return render(request, 'KhoanThuNhapHoc/suaKhoanThuNhaphoc.html', {'form': form,'khoanThuNhapHoc': khoanThuNhapHoc,'cacKhoaHoc': cacKhoaHoc,'cacTenNganh': cacTenNganh,'cacKhoanThu': cacKhoanThu})

@login_required
def XoaKhoanThuNhapHoc(request, id):
    khoanThuNhapHoc = get_object_or_404(KhoanThuNhapHoc, id=id)

    if request.method == 'POST':
        khoanThuNhapHoc.delete()
        alert_content = 'Xóa'
        url = reverse('listKhoanThuNhapHoc') + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'KhoanThuNhapHoc/xoaKhoanThuNhaphoc.html', {'khoanThuNhapHoc': khoanThuNhapHoc})

@login_required
def ListKhoanThuHocPhi(request):
    ListHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    ListNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    ListKhoa = Khoa.objects.all().order_by('-id')
    ListKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    ListLop = Lop.objects.all().order_by('-id')
    if request.method == 'POST':
        loaiBaoCao = request.POST.get('loaiBaoCao')
        if loaiBaoCao == 'tongHop':
            BaoCaoTongHop()

    return render(request, 'KhoanThuHocPhi/listKhoanThu.html', {'ListHeDaoTao': ListHeDaoTao, 'ListNganhDaoTao': ListNganhDaoTao, 'ListKhoa': ListKhoa, 'ListKhoaHoc': ListKhoaHoc, 'ListLop': ListLop})

@login_required
def ListSinhVien(request,id):
   lop = get_object_or_404(Lop, id=id)
   ListSinhVien = SinhVien.objects.filter(lop=id).order_by('-id')
   listKhoanThuHocPhi = KhoanThuHocPhi.objects.filter(mucHocPhi__lop = lop.lop).order_by('-id')
   alertContent = request.GET.get('alertContent')
   return render(request, 'KhoanThuHocPhi/listSinhVien.html', {'listKhoanThuHocPhi': listKhoanThuHocPhi,'ListSinhVien': ListSinhVien, 'lop': lop,'alertContent': alertContent,'lopId':id})

#Batch insert 
@login_required
def AddKhoanThuHocPhi(request,id):
    lop = get_object_or_404(Lop, id=id)
    listSinhVien = SinhVien.objects.filter(lop__id = id).order_by('-id')
    listMucHocPhi= ThietLapMucHocPhi.objects.filter(lop = lop.lop).order_by('-id')
    if request.method == 'POST':
        form = BatchKhoanThuHocPhiForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTienMienGiam'] = data_copy['soTienMienGiam'].replace(",", "") 
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            mucHocPhiId = request.POST.get('mucHocPhi')
            mucHocPhi = get_object_or_404(ThietLapMucHocPhi, id=mucHocPhiId)
            hanNop  = request.POST.get('hanNop')
            noiDungThu  = request.POST.get('noiDungThu')
            ghiChu = request.POST.get('ghiChu')
            maSoThue = request.POST.get('maSoThue')
            soTienMienGiam = request.POST.get('soTienMienGiam').replace(",", "") 
            soTien = request.POST.get('soTien').replace(",", "") 
            for sinhVien in listSinhVien:
                khoanThuHocPhi = KhoanThuHocPhi(sinhVien=sinhVien, mucHocPhi=mucHocPhi, hanNop = hanNop , noiDungThu=noiDungThu, ghiChu=ghiChu,
                                                maSoThue = maSoThue, soTienMienGiam = soTienMienGiam, soTien = soTien, trangThai = 0 )
                khoanThuHocPhi.save()
            alert_content = 'Thêm'
            url = reverse('listSinhVien', args=[id]) + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = BatchKhoanThuHocPhiForm()
    return render(request, 'KhoanThuHocPhi/addKhoanThuHocPhi.html', {'form': form, 'listSinhVien': listSinhVien,'listMucHocPhi': listMucHocPhi,'lopId':id})

@login_required
def NewKhoanThuHocPhi(request,id):
    lop = get_object_or_404(Lop, id=id)
    listSinhVien = SinhVien.objects.filter(lop__id = id).order_by('-id')
    listMucHocPhi= ThietLapMucHocPhi.objects.filter(lop = lop.lop).order_by('-id')
    if request.method == 'POST':
        form = KhoanThuHocPhiForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTienMienGiam'] = data_copy['soTienMienGiam'].replace(",", "") 
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listSinhVien', args=[id]) + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = KhoanThuHocPhiForm()
    return render(request, 'KhoanThuHocPhi/newKhoanThuHocPhi.html', {'form': form, 'listSinhVien': listSinhVien,'listMucHocPhi': listMucHocPhi,'lopId':id})


@login_required
def SuaKhoanThuHocPhi(request, id):
    
    khoanThuHocPhi = KhoanThuHocPhi.objects.get(id=id)
    listMucHocPhi= ThietLapMucHocPhi.objects.filter(lop = khoanThuHocPhi.sinhVien.lop.lop).order_by('-id')
    if request.method == 'POST':
        khoanThuHocPhi.sinhVien = get_object_or_404(SinhVien, id=request.POST.get('sinhVien'))
        khoanThuHocPhi.mucHocPhi = get_object_or_404(ThietLapMucHocPhi, id=request.POST.get('mucHocPhi'))
        khoanThuHocPhi.hanNop = request.POST.get('hanNop')
        khoanThuHocPhi.noiDungThu = request.POST.get('noiDungThu')
        khoanThuHocPhi.ghiChu = request.POST.get('ghiChu')
        khoanThuHocPhi.maSoThue = request.POST.get('maSoThue')
        khoanThuHocPhi.soTienMienGiam = request.POST.get('soTienMienGiam').replace(",", "")
        khoanThuHocPhi.soTien = request.POST.get('soTien').replace(",", "")

        khoanThuHocPhi.save()
        alert_content = 'Sửa'
        url = reverse('listSinhVien', args=[khoanThuHocPhi.sinhVien.lop.id]) + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        
        form = KhoanThuHocPhiForm()

    return render(request, 'KhoanThuHocPhi/suaKhoanThuHocPhi.html', {'form': form,'khoanThuHocPhi': khoanThuHocPhi,'listMucHocPhi': listMucHocPhi})

@login_required
def XoaKhoanThuHocPhi(request, id):
    khoanThuHocPhi = get_object_or_404(KhoanThuHocPhi, id=id)

    if request.method == 'POST':
        khoanThuHocPhi.delete()
        alert_content = 'Xóa'
        url = reverse('listSinhVien', args=[khoanThuHocPhi.sinhVien.lop.id]) + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'KhoanThuHocPhi/xoaKhoanThuHocPhi.html', {'khoanThuHocPhi': khoanThuHocPhi})

@login_required
def XemKhoanThuHocPhi(request, id):
    khoanThuHocPhi = get_object_or_404(KhoanThuHocPhi, id=id)
    tongSoTien = format(khoanThuHocPhi.soTien, ",.0f")
    return render(request, 'KhoanThuHocPhi/xemKhoanThuHocPhi.html', {'khoanThuHocPhi': khoanThuHocPhi})

@login_required
def XemCacKhoanThuCuaSV(request,id):
   sinhVien = get_object_or_404(SinhVien, id=id)
   listKhoanThu = KhoanThuHocPhi.objects.filter(sinhVien=id).order_by('-id')
   alertContent = request.GET.get('alertContent')
   return render(request, 'KhoanThuHocPhi/xemCacKhoanThuCuaSV.html', {'listKhoanThu': listKhoanThu,'sinhVien': sinhVien,'alertContent': alertContent})


@login_required
def ChonBaoCao(request):
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacKhoa = Khoa.objects.all().order_by('-id')
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacNamHoc = NamHoc.objects.all().order_by('-id')
    cacLop = Lop.objects.all().order_by('-id')
    cacLoaiBaoCao = ["Báo cáo tổng hợp","Báo cáo tổng hợp theo lớp","Báo cáo công nợ sinh viên","Báo cáo công nợ sinh viên theo lớp"]

    if request.method == 'POST':
        loaiBaoCao = request.POST.get('loaiBaoCao')
        if loaiBaoCao == 'Báo cáo tổng hợp':
            return BaoCaoTongHop()
        elif loaiBaoCao == 'Báo cáo tổng hợp theo lớp':
            lop = request.POST.get('lop')
            return BaoCaoTongHopTheoLop(lop)
        elif loaiBaoCao == 'Báo cáo công nợ sinh viên':
            return BaoCaoCongNoSinhVien()
        elif loaiBaoCao == 'Báo cáo công nợ sinh viên theo lớp':
            lop = request.POST.get('lop')
            return BaoCaoCongNoTheoLop(lop)

    return render(request, 'BaoCaoThu/chonBaoCao.html', {'cacHeDaoTao': cacHeDaoTao, 'cacNganhDaoTao': cacNganhDaoTao, 'cacKhoa': cacKhoa, 'cacNamHoc': cacNamHoc, 'cacKhoaHoc': cacKhoaHoc, 'cacLop': cacLop, 'cacLoaiBaoCao': cacLoaiBaoCao})

def BaoCaoTongHop():
    ListKhoa = Khoa.objects.all().order_by('-id')
    workbook = Workbook()
    worksheet = workbook.active
    for khoa in ListKhoa:
        queryset = KhoanThuHocPhi.objects.filter(mucHocPhi__khoa=khoa.tenKhoa).order_by('sinhVien__maSV')
        worksheet = workbook.create_sheet(title=khoa.tenKhoa)
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if obj.mucHocPhi.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                obj.mucHocPhi.namHoc,
                hocKy,
                format(obj.mucHocPhi.soTien, ","),
                format(obj.soTienMienGiam, ","),
                format(obj.soTien, ","),
                trangThai

            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = urllib.parse.quote('Báo cáo tổng hợp.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response


def BaoCaoTongHopTheoLop(lopId):
    lopDuocChon = get_object_or_404(Lop, id=lopId)
    listNamHoc = NamHoc.objects.all().order_by('-id')
    name = 'Báo cáo tổng hợp theo lớp '+ lopDuocChon.lop +' .xlsx'
    filename = urllib.parse.quote(name)
    workbook = Workbook()
    worksheet = workbook.active
    for namHoc in listNamHoc:
        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=1) & Q(mucHocPhi__lop=lopDuocChon.lop)).order_by('sinhVien__maSV')
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 1")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if obj.mucHocPhi.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                obj.mucHocPhi.namHoc,
                hocKy,
                format(obj.mucHocPhi.soTien, ","),
                format(obj.soTienMienGiam, ","),
                format(obj.soTien, ","),
                trangThai
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=0) & Q(mucHocPhi__lop=lopDuocChon.lop))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 2")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Năm học','Kỳ học','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ','Trạng thái']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(queryset, 2):
            if obj.trangThai == True:
                trangThai = "Đã hoàn thành"
            else:
                trangThai = "Chưa hoàn thành"

            if obj.mucHocPhi.hocKy == True:
                hocKy = "Học kỳ 1"
            else:
                hocKy = "Học kỳ 2"
            
            row = [
                obj.sinhVien.maSV,
                obj.sinhVien.hoTen,
                obj.sinhVien.ngaySinh,
                obj.sinhVien.lop.lop,
                obj.mucHocPhi.namHoc,
                hocKy,
                format(obj.mucHocPhi.soTien, ","),
                format(obj.soTienMienGiam, ","),
                format(obj.soTien, ","),
                trangThai
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

def BaoCaoCongNoSinhVien():
    ListKhoa = Khoa.objects.all().order_by('-id')
    workbook = Workbook()
    worksheet = workbook.active
    for khoa in ListKhoa:
        queryset = KhoanThuHocPhi.objects.filter(mucHocPhi__khoa=khoa.tenKhoa).filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('mucHocPhi__soTien'))
        worksheet = workbook.create_sheet(title=khoa.tenKhoa)
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']

        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = urllib.parse.quote('Báo cáo công nợ sinh viên.xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

def BaoCaoCongNoTheoLop(lopId):
    lopDuocChon = get_object_or_404(Lop, id=lopId)
    listNamHoc = NamHoc.objects.all().order_by('-id')
    name = 'Báo cáo công nợ sinh viên theo lớp '+ lopDuocChon.lop +' .xlsx'
    filename = urllib.parse.quote(name)
    workbook = Workbook()
    worksheet = workbook.active
    for namHoc in listNamHoc:
        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=1) & Q(mucHocPhi__lop=lopDuocChon.lop)).filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('mucHocPhi__soTien'))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 1")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

        queryset = KhoanThuHocPhi.objects.filter(Q(mucHocPhi__namHoc=namHoc.namHoc) & Q(mucHocPhi__hocKy=0) & Q(mucHocPhi__lop=lopDuocChon.lop)).filter(trangThai=0)
        aggregate_queryset = queryset.values('sinhVien__maSV','sinhVien__hoTen','sinhVien__ngaySinh','sinhVien__lop__lop',).annotate(sum_soTien=Sum('mucHocPhi__soTien'))
        worksheet = workbook.create_sheet(title=namHoc.namHoc + " - Học kỳ 2")
        column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền chưa nộp']
        for col_num, column_name in enumerate(column_names, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_name
            
            # Adjust the column width to auto-stretch
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].bestFit = True

        for row_num, obj in enumerate(aggregate_queryset, 2):
            row = [
                obj['sinhVien__maSV'],
                obj['sinhVien__hoTen'],
                obj['sinhVien__ngaySinh'],
                obj['sinhVien__lop__lop'],
                format(obj['sum_soTien'], ","),
            ]
            for col_num, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
        
        # Adjust the column widths to auto-stretch after populating the data
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.remove(workbook['Sheet'])
    workbook.save(response)
    return response

# def BaoCaoTongHop(self):
#     ListKhoa = Khoa.objects.all().order_by('-id')
#     filename = 'data_export.xlsx'

#     # Create a new workbook and get the active sheet
#     workbook = Workbook()
#     worksheet = workbook.active
#     for khoa in ListKhoa:
#         queryset = KhoanThuHocPhi.objects.filter(mucHocPhi__khoa=khoa.tenKhoa)
#         worksheet = workbook.create_sheet(title=khoa.tenKhoa)
#         # Write column headers
#         column_names = ['Mã sinh viên', 'Họ tên', 'Ngày sinh','Tên lớp','Số tiền phải nộp','Số tiền miễn giảm','Còn nộp trong kỳ']
#         for col_num, column_name in enumerate(column_names, 1):
#             cell = worksheet.cell(row=1, column=col_num)
#             cell.value = column_name

#             column_letter = get_column_letter(col_num)
#             column_dimensions = worksheet.column_dimensions[column_letter]
#             column_dimensions.width = max(column_dimensions.width, len(column_name) + 2)

#         # Write data rows
#         for row_num, obj in enumerate(queryset, 2):
#             row = [
#                 obj.sinhVien.maSV,
#                 obj.sinhVien.hoTen,
#                 obj.sinhVien.ngaySinh,
#                 obj.sinhVien.lop.lop,
#                 format(obj.mucHocPhi.soTien, ","),
#                 format(obj.soTienMienGiam, ","),
#                 format(obj.soTien, ","),
#             ]
#             for col_num, value in enumerate(row, 1):
#                 cell = worksheet.cell(row=row_num, column=col_num)
                
#                 # Remove time zone information from datetime objects
#                 if isinstance(value, datetime):
#                     value = value.replace(tzinfo=None)
                
#                 cell.value = value
#                 column_letter = get_column_letter(col_num)
#                 column_dimensions = worksheet.column_dimensions[column_letter]
#                 column_dimensions.width = max(column_dimensions.width, len(str(value)), len(column_names[col_num-1])) + 2
    
#     # Set response headers for file download
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     workbook.remove(workbook['Sheet'])
#     # Save the workbook to the response
#     workbook.save(response)
#     return response

@login_required
def ListKeHoachThu(request):
   alertContent = request.GET.get('alertContent')
   listKeHoachThu = KeHoachThu.objects.all().order_by('-id')
   return render(request, 'KeHoachThu/listKeHoachThu.html', {'listKeHoachThu': listKeHoachThu, 'alertContent': alertContent})

@login_required
def NewKeHoachThu(request):
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacKhoanThu = DanhMucThuChi.objects.filter(isKhoanThu = True).order_by('-id')
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    if request.method == 'POST':
        form = KeHoachThuForm(request.POST)
        data_copy = request.POST.copy()
        data_copy['soTien'] = data_copy['soTien'].replace(",", "") 
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listKeHoachThu') + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = KeHoachThuForm()
    return render(request, 'KeHoachThu/newKeHoachThu.html', {'form': form, 'cacHeDaoTao': cacHeDaoTao, 'cacKhoanThu': cacKhoanThu,'cacKhoaHoc': cacKhoaHoc,'cacNganhDaoTao': cacNganhDaoTao})

@login_required
def SuaKeHoachThu(request, id):

    keHoachThu = get_object_or_404(KeHoachThu, id=id)
    cacKhoaHoc = KhoaHoc.objects.all().order_by('-id')
    cacNganhDaoTao = NganhDaoTao.objects.all().order_by('-id')
    cacKhoanThu = DanhMucThuChi.objects.filter(isKhoanThu = True).order_by('-id')
    cacHeDaoTao = HeDaoTao.objects.all().order_by('-id')
    if request.method == 'POST':
        keHoachThu.khoaHoc =  get_object_or_404(KhoaHoc, id=request.POST.get('khoaHoc'))
        keHoachThu.khoanThu = get_object_or_404(DanhMucThuChi, id=request.POST.get('khoanThu'))
        keHoachThu.nganhDaoTao = get_object_or_404(NganhDaoTao, id=request.POST.get('nganhDaoTao'))
        keHoachThu.soTien = request.POST.get('soTien').replace(",", "")
        keHoachThu.dotThu = request.POST.get('dotThu')
        keHoachThu.ngayBatDau = request.POST.get('ngayBatDau')
        keHoachThu.ngayKetThuc = request.POST.get('ngayKetThuc')
        keHoachThu.noiDung = request.POST.get('noiDung')
        keHoachThu.save()
        alert_content = 'Sửa'
        url = reverse('listKeHoachThu') + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        keHoachThu = KeHoachThu.objects.get(id=id)
        form = KeHoachThuForm()


    return render(request, 'KeHoachThu/suaKeHoachThu.html', {'form': form,'keHoachThu': keHoachThu,'cacHeDaoTao': cacHeDaoTao,'cacKhoaHoc': cacKhoaHoc,'cacNganhDaoTao': cacNganhDaoTao,'cacKhoanThu': cacKhoanThu})

@login_required
def XoaKeHoachThu(request, id):
    keHoachThu = get_object_or_404(KeHoachThu, id=id)
    soTien = format(keHoachThu.soTien, ",.0f")
    if request.method == 'POST':
        keHoachThu.delete()
        alert_content = 'Xóa'
        url = reverse('listKeHoachThu') + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'KeHoachThu/xoaKeHoachThu.html', {'keHoachThu': keHoachThu, 'soTien': soTien})

@login_required
def XemKeHoachThu(request, id):
    keHoachThu = get_object_or_404(KeHoachThu, id=id)
    soTien = format(keHoachThu.soTien, ",.0f")
    return render(request, 'KeHoachThu/xemKeHoachThu.html', {'keHoachThu': keHoachThu, 'soTien': soTien})


@login_required
def ChonSinhVien(request):
    listSinhVien = SinhVien.objects.all()
    if request.method == 'POST':
            maSV = request.POST.get('sinhVien')
            return redirect('thanhToan', maSV)
    else:
        return render(request, 'ChiTraHocPhi/chonSinhVien.html', {'listSinhVien': listSinhVien})

@login_required
def ThanhToan(request,id):
    sinhVien = get_object_or_404(SinhVien, id=id)
    listKhoanThuHocPhi = KhoanThuHocPhi.objects.filter(sinhVien = sinhVien).order_by('-id')
    if request.method == 'POST':
        form = ChiTraHocPhiForm(request.POST)
        data_copy = request.POST.copy()
        mucHocPhi = data_copy['soTienDaNop'].replace(",", "") 

        data_copy['soTienDaNop'] = data_copy['soTienDaNop'].replace(",", "") 
        form.data = data_copy

        if form.is_valid():
            form.save()
            
            alert_content = 'Thêm'
            url = reverse('xemCacKhoanThuSV', args=[id]) + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = ChiTraHocPhiForm()
    return render(request, 'ChiTraHocPhi/thanhToan.html', {'form': form, 'listKhoanThuHocPhi': listKhoanThuHocPhi, 'sinhVien': sinhVien, 'idSV': id})


@login_required
def ListDanhMucThuChi(request):
   alertContent = request.GET.get('alertContent')
   ListDanhMucThuChi = DanhMucThuChi.objects.all().order_by('-ngayThem')
   return render(request, 'DanhMucThuChi/listDanhMucThuChi.html', {'ListDanhMucThuChi': ListDanhMucThuChi, 'alertContent': alertContent})

@login_required
def NewDanhMucThuChi(request):
    if request.method == 'POST':
        form = DanhMucThuChiForm(request.POST)
        data_copy = request.POST.copy()
        loaiKhoan = request.POST.get('loaiKhoan')
        print(loaiKhoan)
        if(int(loaiKhoan) == 1):
            data_copy['isKhoanThu'] = True
            data_copy['isKhoanChi'] = False
        else:
            data_copy['isKhoanThu'] = False
            data_copy['isKhoanChi']  = True 
        data_copy['isKhoanThuNhapHoc'] = request.POST.get('isKhoanThuNhapHoc') 
        data_copy['isHocPhi'] = request.POST.get('isHocPhi') 
        data_copy['isThuTrucTuyen'] = request.POST.get('isThuTrucTuyen') 
        form.data = data_copy
        if form.is_valid():
            form.save()
            alert_content = 'Thêm'
            url = reverse('listDanhMucThuChi') + f'?alertContent={alert_content}'
            return redirect(url)
    else:
        form = DanhMucThuChiForm()
    return render(request, 'DanhMucThuChi/newDanhMucThuChi.html', {'form': form})

@login_required
def SuaDanhMucThuChi(request, id):
    danhMucThuChi = get_object_or_404(DanhMucThuChi, id=id)
    
    if request.method == 'POST':
        danhMucThuChi.tenKhoan = request.POST.get('tenKhoan')
        loaiKhoan = request.POST.get('loaiKhoan')
        if(int(loaiKhoan) == 1):
            danhMucThuChi.isKhoanThu = 1
            danhMucThuChi.isKhoanChi = 0
        else:
            danhMucThuChi.isKhoanThu = 0
            danhMucThuChi.isKhoanChi = 1

        danhMucThuChi.isKhoanThuNhapHoc = request.POST.get('isKhoanThuNhapHoc')
        danhMucThuChi.isHocPhi = request.POST.get('isHocPhi')
        danhMucThuChi.isThuTrucTuyen = request.POST.get('isThuTrucTuyen')
        danhMucThuChi.save()

        alert_content = 'Sửa'
        url = reverse('listDanhMucThuChi') + f'?alertContent={alert_content}'
        return redirect(url)
    else:
        danhMucThuChi = DanhMucThuChi.objects.get(id=id)
        form = DanhMucThuChiForm()

    return render(request, 'DanhMucThuChi/suaDanhMucThuChi.html', {'form': form,'danhMucThuChi': danhMucThuChi})

@login_required
def XoaDanhMucThuChi(request, id):
    danhMucThuChi = get_object_or_404(DanhMucThuChi, id=id)

    if request.method == 'POST':
        danhMucThuChi.delete()
        alert_content = 'Xóa'
        url = reverse('listDanhMucThuChi') + f'?alertContent={alert_content}'
        return redirect(url)

    return render(request, 'DanhMucThuChi/xoaDanhMucThuChi.html', {'danhMucThuChi': danhMucThuChi})


def index(request):
    return render(request, "payment/index.html", {"title": "Danh sách demo"})


def hmacsha512(key, data):
    byteKey = key.encode('utf-8')
    byteData = data.encode('utf-8')
    return hmac.new(byteKey, byteData, hashlib.sha512).hexdigest()


def payment(request):

    if request.method == 'POST':
        # Process input data and build url payment
        form = PaymentForm(request.POST)
        if form.is_valid():
            order_type = form.cleaned_data['order_type']
            order_id = form.cleaned_data['order_id']
            amount = form.cleaned_data['amount']
            order_desc = form.cleaned_data['order_desc']
            bank_code = form.cleaned_data['bank_code']
            language = form.cleaned_data['language']
            ipaddr = get_client_ip(request)
            # Build URL Payment
            vnp = vnpay()
            vnp.requestData['vnp_Version'] = '2.1.0'
            vnp.requestData['vnp_Command'] = 'pay'
            vnp.requestData['vnp_TmnCode'] = settings.VNPAY_TMN_CODE
            vnp.requestData['vnp_Amount'] = amount * 100
            vnp.requestData['vnp_CurrCode'] = 'VND'
            vnp.requestData['vnp_TxnRef'] = order_id
            vnp.requestData['vnp_OrderInfo'] = order_desc
            vnp.requestData['vnp_OrderType'] = order_type
            # Check language, default: vn
            if language and language != '':
                vnp.requestData['vnp_Locale'] = language
            else:
                vnp.requestData['vnp_Locale'] = 'vn'
                # Check bank_code, if bank_code is empty, customer will be selected bank on VNPAY
            if bank_code and bank_code != "":
                vnp.requestData['vnp_BankCode'] = bank_code

            vnp.requestData['vnp_CreateDate'] = datetime.now().strftime('%Y%m%d%H%M%S')  # 20150410063022
            vnp.requestData['vnp_IpAddr'] = ipaddr
            vnp.requestData['vnp_ReturnUrl'] = settings.VNPAY_RETURN_URL
            vnpay_payment_url = vnp.get_payment_url(settings.VNPAY_PAYMENT_URL, settings.VNPAY_HASH_SECRET_KEY)
            print(vnpay_payment_url)
            return redirect(vnpay_payment_url)
        else:
            print("Form input not validate")
    else:
        return render(request, "payment/payment.html", {"title": "Thanh toán"})


def payment_ipn(request):
    inputData = request.GET
    if inputData:
        vnp = vnpay()
        vnp.responseData = inputData.dict()
        order_id = inputData['vnp_TxnRef']
        amount = inputData['vnp_Amount']
        order_desc = inputData['vnp_OrderInfo']
        vnp_TransactionNo = inputData['vnp_TransactionNo']
        vnp_ResponseCode = inputData['vnp_ResponseCode']
        vnp_TmnCode = inputData['vnp_TmnCode']
        vnp_PayDate = inputData['vnp_PayDate']
        vnp_BankCode = inputData['vnp_BankCode']
        vnp_CardType = inputData['vnp_CardType']
        if vnp.validate_response(settings.VNPAY_HASH_SECRET_KEY):
            # Check & Update Order Status in your Database
            # Your code here
            firstTimeUpdate = True
            totalamount = True
            if totalamount:
                if firstTimeUpdate:
                    if vnp_ResponseCode == '00':
                        print('Payment Success. Your code implement here')
                    else:
                        print('Payment Error. Your code implement here')

                    # Return VNPAY: Merchant update success
                    result = JsonResponse({'RspCode': '00', 'Message': 'Confirm Success'})
                else:
                    # Already Update
                    result = JsonResponse({'RspCode': '02', 'Message': 'Order Already Update'})
            else:
                # invalid amount
                result = JsonResponse({'RspCode': '04', 'Message': 'invalid amount'})
        else:
            # Invalid Signature
            result = JsonResponse({'RspCode': '97', 'Message': 'Invalid Signature'})
    else:
        result = JsonResponse({'RspCode': '99', 'Message': 'Invalid request'})

    return result


def payment_return(request):
    inputData = request.GET
    if inputData:
        vnp = vnpay()
        vnp.responseData = inputData.dict()
        order_id = inputData['vnp_TxnRef']
        amount = int(inputData['vnp_Amount']) / 100
        order_desc = inputData['vnp_OrderInfo']
        vnp_TransactionNo = inputData['vnp_TransactionNo']
        vnp_ResponseCode = inputData['vnp_ResponseCode']
        vnp_TmnCode = inputData['vnp_TmnCode']
        vnp_PayDate = inputData['vnp_PayDate']
        vnp_BankCode = inputData['vnp_BankCode']
        vnp_CardType = inputData['vnp_CardType']

        payment = Payment_VNPay.objects.create(
            order_id = order_id,
            amount = amount,
            order_desc = order_desc,
            vnp_TransactionNo = vnp_TransactionNo,
            vnp_ResponseCode = vnp_ResponseCode
        )


        if vnp.validate_response(settings.VNPAY_HASH_SECRET_KEY):
            if vnp_ResponseCode == "00":
                return render(request, "payment/payment_return.html", {"title": "Kết quả thanh toán",
                                                               "result": "Thành công", "order_id": order_id,
                                                               "amount": amount,
                                                               "order_desc": order_desc,
                                                               "vnp_TransactionNo": vnp_TransactionNo,
                                                               "vnp_ResponseCode": vnp_ResponseCode})
            else:
                return render(request, "payment/payment_return.html", {"title": "Kết quả thanh toán",
                                                               "result": "Lỗi", "order_id": order_id,
                                                               "amount": amount,
                                                               "order_desc": order_desc,
                                                               "vnp_TransactionNo": vnp_TransactionNo,
                                                               "vnp_ResponseCode": vnp_ResponseCode})
        else:
            return render(request, "payment/payment_return.html",
                          {"title": "Kết quả thanh toán", "result": "Lỗi", "order_id": order_id, "amount": amount,
                           "order_desc": order_desc, "vnp_TransactionNo": vnp_TransactionNo,
                           "vnp_ResponseCode": vnp_ResponseCode, "msg": "Sai checksum"})
    else:
        return render(request, "payment/payment_return.html", {"title": "Kết quả thanh toán", "result": ""})


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

n = random.randint(10**11, 10**12 - 1)
n_str = str(n)
while len(n_str) < 12:
    n_str = '0' + n_str


def query(request):
    if request.method == 'GET':
        return render(request, "payment/query.html", {"title": "Kiểm tra kết quả giao dịch"})

    url = settings.VNPAY_API_URL
    secret_key = settings.VNPAY_HASH_SECRET_KEY
    vnp_TmnCode = settings.VNPAY_TMN_CODE
    vnp_Version = '2.1.0'

    vnp_RequestId = n_str
    vnp_Command = 'querydr'
    vnp_TxnRef = request.POST['order_id']
    vnp_OrderInfo = 'kiem tra gd'
    vnp_TransactionDate = request.POST['trans_date']
    vnp_CreateDate = datetime.now().strftime('%Y%m%d%H%M%S')
    vnp_IpAddr = get_client_ip(request)

    hash_data = "|".join([
        vnp_RequestId, vnp_Version, vnp_Command, vnp_TmnCode,
        vnp_TxnRef, vnp_TransactionDate, vnp_CreateDate,
        vnp_IpAddr, vnp_OrderInfo
    ])

    secure_hash = hmac.new(secret_key.encode(), hash_data.encode(), hashlib.sha512).hexdigest()

    data = {
        "vnp_RequestId": vnp_RequestId,
        "vnp_TmnCode": vnp_TmnCode,
        "vnp_Command": vnp_Command,
        "vnp_TxnRef": vnp_TxnRef,
        "vnp_OrderInfo": vnp_OrderInfo,
        "vnp_TransactionDate": vnp_TransactionDate,
        "vnp_CreateDate": vnp_CreateDate,
        "vnp_IpAddr": vnp_IpAddr,
        "vnp_Version": vnp_Version,
        "vnp_SecureHash": secure_hash
    }

    headers = {"Content-Type": "application/json"}

    response = requests.post(url, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        response_json = json.loads(response.text)
    else:
        response_json = {"error": f"Request failed with status code: {response.status_code}"}

    return render(request, "payment/query.html", {"title": "Kiểm tra kết quả giao dịch", "response_json": response_json})

def refund(request):
    if request.method == 'GET':
        return render(request, "payment/refund.html", {"title": "Hoàn tiền giao dịch"})

    url = settings.VNPAY_API_URL
    secret_key = settings.VNPAY_HASH_SECRET_KEY
    vnp_TmnCode = settings.VNPAY_TMN_CODE
    vnp_RequestId = n_str
    vnp_Version = '2.1.0'
    vnp_Command = 'refund'
    vnp_TransactionType = request.POST['TransactionType']
    vnp_TxnRef = request.POST['order_id']
    vnp_Amount = request.POST['amount']
    vnp_OrderInfo = request.POST['order_desc']
    vnp_TransactionNo = '0'
    vnp_TransactionDate = request.POST['trans_date']
    vnp_CreateDate = datetime.now().strftime('%Y%m%d%H%M%S')
    vnp_CreateBy = 'user01'
    vnp_IpAddr = get_client_ip(request)

    hash_data = "|".join([
        vnp_RequestId, vnp_Version, vnp_Command, vnp_TmnCode, vnp_TransactionType, vnp_TxnRef,
        vnp_Amount, vnp_TransactionNo, vnp_TransactionDate, vnp_CreateBy, vnp_CreateDate,
        vnp_IpAddr, vnp_OrderInfo
    ])

    secure_hash = hmac.new(secret_key.encode(), hash_data.encode(), hashlib.sha512).hexdigest()

    data = {
        "vnp_RequestId": vnp_RequestId,
        "vnp_TmnCode": vnp_TmnCode,
        "vnp_Command": vnp_Command,
        "vnp_TxnRef": vnp_TxnRef,
        "vnp_Amount": vnp_Amount,
        "vnp_OrderInfo": vnp_OrderInfo,
        "vnp_TransactionDate": vnp_TransactionDate,
        "vnp_CreateDate": vnp_CreateDate,
        "vnp_IpAddr": vnp_IpAddr,
        "vnp_TransactionType": vnp_TransactionType,
        "vnp_TransactionNo": vnp_TransactionNo,
        "vnp_CreateBy": vnp_CreateBy,
        "vnp_Version": vnp_Version,
        "vnp_SecureHash": secure_hash
    }

    headers = {"Content-Type": "application/json"}

    response = requests.post(url, headers=headers, data=json.dumps(data))

    if response.status_code == 200:
        response_json = json.loads(response.text)
    else:
        response_json = {"error": f"Request failed with status code: {response.status_code}"}

    return render(request, "payment/refund.html", {"title": "Kết quả hoàn tiền giao dịch", "response_json": response_json})